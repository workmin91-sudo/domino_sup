from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash, session
import sqlite3
import pandas as pd
from datetime import datetime
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
from functools import wraps
from config import EMAIL_SENDER, EMAIL_PASSWORD, SMTP_SERVER, SMTP_PORT, DATABASE, DEFAULT_MIN_STOCK, LOGIN_PASSWORD

app = Flask(__name__)
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'your-secret-key-here-change-in-production')  # 환경 변수에서 가져옴

def login_required(f):
    """로그인 필요 데코레이터"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session or not session['logged_in']:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def init_db():
    """데이터베이스 초기화"""
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    
    # 재고 테이블
    c.execute('''
        CREATE TABLE IF NOT EXISTS inventory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_name TEXT NOT NULL,
            specification TEXT,
            current_stock INTEGER NOT NULL,
            min_stock INTEGER NOT NULL,
            shortage_quantity INTEGER DEFAULT 0,
            reorder_quantity INTEGER DEFAULT 0,
            status TEXT DEFAULT '정상',
            unit TEXT DEFAULT '개',
            supplier_email TEXT,
            supplier_name TEXT,
            notification_message TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 새 컬럼 추가 (기존 테이블에 컬럼이 없는 경우)
    new_columns = [
        ('specification', 'TEXT'),
        ('shortage_quantity', 'INTEGER DEFAULT 0'),
        ('reorder_quantity', 'INTEGER DEFAULT 0'),
        ('status', 'TEXT DEFAULT "정상"'),
        ('notification_message', 'TEXT')
    ]
    for col_name, col_type in new_columns:
        try:
            c.execute(f'ALTER TABLE inventory ADD COLUMN {col_name} {col_type}')
        except sqlite3.OperationalError:
            pass  # 컬럼이 이미 존재하는 경우
    
    # 발주 내역 테이블
    c.execute('''
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER,
            item_name TEXT NOT NULL,
            order_quantity INTEGER NOT NULL,
            order_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            email_sent INTEGER DEFAULT 0,
            email_sent_at TIMESTAMP,
            FOREIGN KEY (item_id) REFERENCES inventory (id)
        )
    ''')
    
    conn.commit()
    conn.close()

def get_db_connection():
    """데이터베이스 연결"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def load_excel_data():
    """엑셀 파일에서 기본 데이터 자동 로드"""
    try:
        import openpyxl
        
        excel_path = 'sup/domino_inventory_training.xlsx'
        if not os.path.exists(excel_path):
            print(f"엑셀 파일을 찾을 수 없습니다: {excel_path}")
            return
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # 헤더 행 찾기 (재료명, 규격 등이 있는 행)
        header_row = None
        header_map = {}
        for row_idx in range(1, min(15, ws.max_row + 1)):
            row = [str(cell.value) if cell.value else '' for cell in ws[row_idx]]
            row_str = ' '.join(row).lower()
            
            # 헤더 키워드 찾기
            if '재료명' in row_str or '품목명' in row_str:
                header_row = row_idx
                # 컬럼 매핑
                for col_idx, cell_value in enumerate(row):
                    cell_str = str(cell_value).strip()
                    if '재료명' in cell_str or '품목명' in cell_str:
                        header_map['item_name'] = col_idx
                    elif '규격' in cell_str:
                        header_map['specification'] = col_idx
                    elif '단위' in cell_str:
                        header_map['unit'] = col_idx
                    elif '안전재고' in cell_str or '최소재고' in cell_str:
                        header_map['min_stock'] = col_idx
                    elif '현재재고' in cell_str or '현재 재고' in cell_str:
                        header_map['current_stock'] = col_idx
                    elif '부족수량' in cell_str or '부족 수량' in cell_str:
                        header_map['shortage_quantity'] = col_idx
                    elif '발주권장수량' in cell_str or '발주권장' in cell_str or '발수 주량' in cell_str:
                        header_map['reorder_quantity'] = col_idx
                    elif '상태' in cell_str:
                        header_map['status'] = col_idx
                    elif '거래처이메일' in cell_str or '이메일' in cell_str:
                        header_map['supplier_email'] = col_idx
                    elif '담당자알림메시지' in cell_str or '알림메시지' in cell_str:
                        header_map['notification_message'] = col_idx
                break
        
        # 헤더를 찾지 못한 경우 기본 구조 사용 (Row 11부터 데이터)
        if header_row is None:
            header_row = 10
            header_map = {
                'item_name': 1,
                'reorder_quantity': 3
            }
        
        conn = get_db_connection()
        count = 0
        
        # 데이터 행 시작 (헤더 다음 행)
        data_start_row = header_row + 1
        for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, values_only=True):
            # 재료명 확인
            item_name_idx = header_map.get('item_name', 1)
            if not row[item_name_idx] or str(row[item_name_idx]).strip() == '':
                continue
                
            item_name = str(row[item_name_idx]).strip()
            if item_name == 'None' or not item_name:
                continue
            
            # 각 컬럼 읽기
            specification = str(row[header_map.get('specification', 2)]).strip() if header_map.get('specification') and len(row) > header_map.get('specification', 0) and row[header_map.get('specification', 2)] else ''
            unit = str(row[header_map.get('unit', 2)]).strip() if header_map.get('unit') and len(row) > header_map.get('unit', 0) and row[header_map.get('unit', 2)] else '개'
            min_stock = int(float(row[header_map.get('min_stock', 2)])) if header_map.get('min_stock') and len(row) > header_map.get('min_stock', 0) and row[header_map.get('min_stock', 2)] else 10
            current_stock = int(float(row[header_map.get('current_stock', 2)])) if header_map.get('current_stock') and len(row) > header_map.get('current_stock', 0) and row[header_map.get('current_stock', 2)] else 0
            shortage_quantity = int(float(row[header_map.get('shortage_quantity', 2)])) if header_map.get('shortage_quantity') and len(row) > header_map.get('shortage_quantity', 0) and row[header_map.get('shortage_quantity', 2)] else 0
            reorder_quantity = int(float(row[header_map.get('reorder_quantity', 3)])) if header_map.get('reorder_quantity') and len(row) > header_map.get('reorder_quantity', 0) and row[header_map.get('reorder_quantity', 3)] else 0
            status = str(row[header_map.get('status', 2)]).strip() if header_map.get('status') and len(row) > header_map.get('status', 0) and row[header_map.get('status', 2)] else '정상'
            supplier_email = str(row[header_map.get('supplier_email', 2)]).strip() if header_map.get('supplier_email') and len(row) > header_map.get('supplier_email', 0) and row[header_map.get('supplier_email', 2)] else ''
            notification_message = str(row[header_map.get('notification_message', 2)]).strip() if header_map.get('notification_message') and len(row) > header_map.get('notification_message', 0) and row[header_map.get('notification_message', 2)] else ''
            
            # 부족수량 계산 (안전재고 - 현재재고)
            if shortage_quantity == 0 and current_stock < min_stock:
                shortage_quantity = min_stock - current_stock
            
            # 상태 계산
            if status == '정상' or not status:
                if current_stock <= min_stock:
                    status = '부족'
                else:
                    status = '정상'
            
            # 기존 항목 확인
            existing = conn.execute('SELECT id FROM inventory WHERE item_name = ?', (item_name,)).fetchone()
            
            if existing:
                # 기존 항목 업데이트
                conn.execute('''
                    UPDATE inventory 
                    SET specification = ?, unit = ?, min_stock = ?, shortage_quantity = ?, 
                        reorder_quantity = ?, status = ?, supplier_email = ?, notification_message = ?, updated_at = ?
                    WHERE item_name = ?
                ''', (specification, unit, min_stock, shortage_quantity, reorder_quantity, status, supplier_email, notification_message, datetime.now(), item_name))
            else:
                # 새 항목 추가
                conn.execute('''
                    INSERT INTO inventory (item_name, specification, current_stock, min_stock, shortage_quantity, 
                                         reorder_quantity, status, unit, supplier_email, notification_message)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (item_name, specification, current_stock, min_stock, shortage_quantity, reorder_quantity, status, unit, supplier_email, notification_message))
                count += 1
        
        conn.commit()
        conn.close()
        
        if count > 0:
            print(f"엑셀 파일에서 {count}개의 새 항목이 로드되었습니다.")
    except Exception as e:
        print(f"엑셀 파일 로드 중 오류: {str(e)}")
        import traceback
        traceback.print_exc()

def check_low_stock():
    """재고 부족 항목 확인"""
    conn = get_db_connection()
    low_stock_items = conn.execute('''
        SELECT * FROM inventory 
        WHERE current_stock <= min_stock
    ''').fetchall()
    conn.close()
    return low_stock_items

def send_order_email(item_name, current_stock, min_stock, supplier_email, supplier_name):
    """발주서 이메일 발송"""
    try:
        # 이메일 내용 생성
        msg = MIMEMultipart()
        msg['From'] = EMAIL_SENDER
        msg['To'] = supplier_email
        msg['Subject'] = f'[발주요청] {item_name} 재고 부족 안내'
        
        # 발수 주량이 설정되어 있으면 사용, 없으면 기본 계산
        conn_check = get_db_connection()
        item_data = conn_check.execute('SELECT reorder_quantity FROM inventory WHERE item_name = ?', (item_name,)).fetchone()
        conn_check.close()
        
        if item_data and item_data['reorder_quantity'] and item_data['reorder_quantity'] > 0:
            order_quantity = item_data['reorder_quantity']
        else:
            order_quantity = min_stock * 3 - current_stock  # 최소 재고의 3배까지 보충
        
        body = f"""
안녕하세요 {supplier_name}님,

재고 관리 시스템에서 다음과 같은 발주 요청이 발생했습니다.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
발주 정보
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
품목명: {item_name}
현재 재고: {current_stock}
최소 재고 기준: {min_stock}
발주 수량: {order_quantity}
발주 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

위 품목의 재고가 부족하여 발주를 요청드립니다.
빠른 시일 내에 배송 부탁드립니다.

감사합니다.
        """
        
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # 이메일 발송
        if EMAIL_PASSWORD:
            server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.send_message(msg)
            server.quit()
            return True
        else:
            print(f"[이메일 미발송] 비밀번호 미설정: {item_name} -> {supplier_email}")
            return False
    except Exception as e:
        print(f"이메일 발송 오류: {e}")
        return False

@app.route('/login', methods=['GET', 'POST'])
def login():
    """로그인 페이지"""
    if request.method == 'POST':
        password = request.form.get('password', '')
        if password == LOGIN_PASSWORD:
            session['logged_in'] = True
            flash('로그인 성공!', 'success')
            return redirect(url_for('index'))
        else:
            flash('비밀번호가 올바르지 않습니다.', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    """로그아웃"""
    session.pop('logged_in', None)
    flash('로그아웃되었습니다.', 'success')
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    """메인 페이지"""
    conn = get_db_connection()
    items = conn.execute('''
        SELECT * FROM inventory ORDER BY updated_at DESC
    ''').fetchall()
    
    # 부족수량 및 상태 자동 계산 및 업데이트
    for item in items:
        shortage_quantity = max(0, item['min_stock'] - item['current_stock']) if item['current_stock'] < item['min_stock'] else 0
        status = '부족' if item['current_stock'] <= item['min_stock'] else '정상'
        
        # 부족수량이나 상태가 변경되었으면 업데이트
        try:
            current_shortage = item['shortage_quantity'] if item['shortage_quantity'] is not None else 0
        except (KeyError, IndexError):
            current_shortage = 0
        try:
            current_status = item['status'] if item['status'] else '정상'
        except (KeyError, IndexError):
            current_status = '정상'
        
        if current_shortage != shortage_quantity or current_status != status:
            conn.execute('''
                UPDATE inventory 
                SET shortage_quantity = ?, status = ?, updated_at = ?
                WHERE id = ?
            ''', (shortage_quantity, status, datetime.now(), item['id']))
    
    conn.commit()
    
    # 업데이트된 데이터 다시 가져오기
    items = conn.execute('''
        SELECT * FROM inventory ORDER BY updated_at DESC
    ''').fetchall()
    conn.close()
    
    # 재고 부족 항목 확인
    low_stock_items = check_low_stock()
    
    return render_template('index.html', items=items, low_stock_items=low_stock_items)

@app.route('/add', methods=['GET', 'POST'])
@login_required
def add_item():
    """재고 항목 추가"""
    if request.method == 'POST':
        item_name = request.form['item_name']
        current_stock = int(request.form['current_stock'])
        min_stock = int(request.form['min_stock'])
        reorder_quantity = int(request.form.get('reorder_quantity', 0))
        unit = request.form.get('unit', '개')
        supplier_email = request.form.get('supplier_email', '')
        supplier_name = request.form.get('supplier_name', '')
        
        conn = get_db_connection()
        conn.execute('''
            INSERT INTO inventory (item_name, current_stock, min_stock, reorder_quantity, unit, supplier_email, supplier_name)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (item_name, current_stock, min_stock, reorder_quantity, unit, supplier_email, supplier_name))
        conn.commit()
        conn.close()
        
        # 재고 부족 확인 및 발주 처리
        if current_stock <= min_stock and supplier_email:
            send_order_email(item_name, current_stock, min_stock, supplier_email, supplier_name)
            # 발주 내역 저장
            conn = get_db_connection()
            order_qty = reorder_quantity if reorder_quantity > 0 else (min_stock * 3 - current_stock)
            conn.execute('''
                INSERT INTO orders (item_id, item_name, order_quantity, email_sent, email_sent_at)
                VALUES ((SELECT id FROM inventory WHERE item_name = ?), ?, ?, 1, ?)
            ''', (item_name, item_name, order_qty, datetime.now()))
            conn.commit()
            conn.close()
            flash(f'{item_name} 재고가 부족하여 발주서가 전송되었습니다.', 'warning')
        else:
            flash('재고 항목이 추가되었습니다.', 'success')
        
        return redirect(url_for('index'))
    
    return render_template('add_item.html')

@app.route('/update/<int:item_id>', methods=['POST'])
@login_required
def update_stock(item_id):
    """재고 수정"""
    new_stock = int(request.form['new_stock'])
    
    conn = get_db_connection()
    item = conn.execute('SELECT * FROM inventory WHERE id = ?', (item_id,)).fetchone()
    
    if item:
        conn.execute('''
            UPDATE inventory 
            SET current_stock = ?, updated_at = ?
            WHERE id = ?
        ''', (new_stock, datetime.now(), item_id))
        conn.commit()
        
        # 재고 부족 확인 및 발주 처리
        if new_stock <= item['min_stock'] and item['supplier_email']:
            send_order_email(
                item['item_name'], 
                new_stock, 
                item['min_stock'], 
                item['supplier_email'],
                item['supplier_name']
            )
            # 발주 내역 저장 - 발수 주량 사용
            try:
                reorder_qty = item['reorder_quantity'] if item['reorder_quantity'] is not None and item['reorder_quantity'] > 0 else 0
            except (KeyError, IndexError):
                reorder_qty = 0
            if reorder_qty > 0:
                order_quantity = reorder_qty
            else:
                order_quantity = item['min_stock'] * 3 - new_stock
            conn.execute('''
                INSERT INTO orders (item_id, item_name, order_quantity, email_sent, email_sent_at)
                VALUES (?, ?, ?, 1, ?)
            ''', (item_id, item['item_name'], order_quantity, datetime.now()))
            conn.commit()
            flash(f'{item["item_name"]} 재고가 부족하여 발주서가 전송되었습니다.', 'warning')
        else:
            flash('재고가 업데이트되었습니다.', 'success')
    
    conn.close()
    return redirect(url_for('index'))

@app.route('/adjust/<int:item_id>', methods=['POST'])
@login_required
def adjust_stock(item_id):
    """재고 수량 빠른 조정 (+/-)"""
    data = request.get_json()
    adjustment = int(data.get('adjustment', 0))
    
    conn = get_db_connection()
    item = conn.execute('SELECT * FROM inventory WHERE id = ?', (item_id,)).fetchone()
    
    if item:
        new_stock = max(0, item['current_stock'] + adjustment)  # 음수 방지
        
        # 부족수량 및 상태 계산
        shortage_quantity = max(0, item['min_stock'] - new_stock) if new_stock < item['min_stock'] else 0
        status = '부족' if new_stock <= item['min_stock'] else '정상'
        
        conn.execute('''
            UPDATE inventory 
            SET current_stock = ?, shortage_quantity = ?, status = ?, updated_at = ?
            WHERE id = ?
        ''', (new_stock, shortage_quantity, status, datetime.now(), item_id))
        conn.commit()
        
        # 재고 부족 확인 및 발주 처리
        email_sent = False
        if new_stock <= item['min_stock'] and item['supplier_email']:
            send_order_email(
                item['item_name'], 
                new_stock, 
                item['min_stock'], 
                item['supplier_email'],
                item['supplier_name']
            )
            # 발주 내역 저장 - 발수 주량 사용
            try:
                reorder_qty = item['reorder_quantity'] if item['reorder_quantity'] is not None and item['reorder_quantity'] > 0 else 0
            except (KeyError, IndexError):
                reorder_qty = 0
            if reorder_qty > 0:
                order_quantity = reorder_qty
            else:
                order_quantity = item['min_stock'] * 3 - new_stock
            conn.execute('''
                INSERT INTO orders (item_id, item_name, order_quantity, email_sent, email_sent_at)
                VALUES (?, ?, ?, 1, ?)
            ''', (item_id, item['item_name'], order_quantity, datetime.now()))
            conn.commit()
            email_sent = True
        
        conn.close()
        return jsonify({
            'success': True,
            'new_stock': new_stock,
            'email_sent': email_sent,
            'message': f'{item["item_name"]} 재고가 {new_stock}로 변경되었습니다.' + (' (발주서 전송됨)' if email_sent else '')
        })
    
    conn.close()
    return jsonify({'success': False, 'message': '항목을 찾을 수 없습니다.'}), 404

@app.route('/update_reorder/<int:item_id>', methods=['POST'])
@login_required
def update_reorder_quantity(item_id):
    """발수 주량 업데이트"""
    data = request.get_json()
    reorder_quantity = int(data.get('reorder_quantity', 0))
    
    conn = get_db_connection()
    item = conn.execute('SELECT * FROM inventory WHERE id = ?', (item_id,)).fetchone()
    
    if item:
        conn.execute('''
            UPDATE inventory 
            SET reorder_quantity = ?, updated_at = ?
            WHERE id = ?
        ''', (reorder_quantity, datetime.now(), item_id))
        conn.commit()
        conn.close()
        return jsonify({
            'success': True,
            'message': f'{item["item_name"]}의 발수 주량이 {reorder_quantity}로 변경되었습니다.'
        })
    
    conn.close()
    return jsonify({'success': False, 'message': '항목을 찾을 수 없습니다.'}), 404

@app.route('/import_excel', methods=['POST'])
@login_required
def import_excel():
    """엑셀 파일에서 데이터 가져오기 (수동)"""
    try:
        count = load_excel_data_manual()
        flash(f'{count}개의 항목이 엑셀 파일에서 가져와졌습니다.', 'success')
        return jsonify({'success': True, 'count': count, 'message': f'{count}개의 항목이 가져와졌습니다.'})
    except Exception as e:
        flash(f'엑셀 파일 가져오기 중 오류: {str(e)}', 'error')
        return jsonify({'success': False, 'message': f'오류: {str(e)}'}), 500

def load_excel_data_manual():
    """엑셀 파일에서 데이터 가져오기 (수동 호출용)"""
    import openpyxl
    
    excel_path = 'sup/domino_inventory_training.xlsx'
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    conn = get_db_connection()
    count = 0
    
    # Row 11부터 데이터 시작 (Row 10은 헤더)
    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True):
        if not row[0] or not row[1]:  # 번호나 품목명이 없으면 스킵
            continue
            
        item_name = str(row[1]).strip()
        if not item_name or item_name == 'None':
            continue
        
        # 발수 주량 (컬럼 3, 인덱스 3)
        reorder_quantity = 0
        if row[3] and str(row[3]).strip():
            try:
                reorder_quantity = int(float(str(row[3])))
            except:
                pass
        
        # 기본값 설정
        current_stock = 0
        min_stock = 10
        
        # 기존 항목 확인
        existing = conn.execute('SELECT id FROM inventory WHERE item_name = ?', (item_name,)).fetchone()
        
        if existing:
            # 기존 항목 업데이트 (발수 주량만 업데이트)
            conn.execute('''
                UPDATE inventory 
                SET reorder_quantity = ?, updated_at = ?
                WHERE item_name = ?
            ''', (reorder_quantity, datetime.now(), item_name))
        else:
            # 새 항목 추가
            conn.execute('''
                INSERT INTO inventory (item_name, current_stock, min_stock, reorder_quantity, unit)
                VALUES (?, ?, ?, ?, ?)
            ''', (item_name, current_stock, min_stock, reorder_quantity, '개'))
            count += 1
    
    conn.commit()
    conn.close()
    return count

@app.route('/place_order/<int:item_id>', methods=['POST'])
@login_required
def place_order(item_id):
    """발주 처리 (수량 선택 가능)"""
    data = request.get_json()
    order_quantity = int(data.get('order_quantity', 0))
    
    if order_quantity <= 0:
        return jsonify({'success': False, 'message': '발주 수량을 입력해주세요.'}), 400
    
    conn = get_db_connection()
    item = conn.execute('SELECT * FROM inventory WHERE id = ?', (item_id,)).fetchone()
    
    if not item:
        conn.close()
        return jsonify({'success': False, 'message': '항목을 찾을 수 없습니다.'}), 404
    
    if not item['supplier_email']:
        conn.close()
        return jsonify({'success': False, 'message': '공급업체 이메일이 설정되지 않았습니다.'}), 400
    
    # 이메일 발송
    email_sent = send_order_email_manual(
        item['item_name'],
        item['current_stock'],
        item['min_stock'],
        item['supplier_email'],
        item['supplier_name'],
        order_quantity
    )
    
    # 발주 내역 저장
    conn.execute('''
        INSERT INTO orders (item_id, item_name, order_quantity, email_sent, email_sent_at)
        VALUES (?, ?, ?, ?, ?)
    ''', (item_id, item['item_name'], order_quantity, 1 if email_sent else 0, datetime.now()))
    conn.commit()
    conn.close()
    
    if email_sent:
        return jsonify({
            'success': True,
            'message': f'{item["item_name"]} {order_quantity}개 발주서가 전송되었습니다.'
        })
    else:
        return jsonify({
            'success': True,
            'message': f'{item["item_name"]} {order_quantity}개 발주 내역이 저장되었습니다. (이메일 미발송)'
        })

def send_order_email_manual(item_name, current_stock, min_stock, supplier_email, supplier_name, order_quantity):
    """발주서 이메일 발송 (수동 발주용)"""
    try:
        # 이메일 내용 생성
        msg = MIMEMultipart()
        msg['From'] = EMAIL_SENDER
        msg['To'] = supplier_email
        msg['Subject'] = f'[발주요청] {item_name} 발주 요청'
        
        body = f"""
안녕하세요 {supplier_name}님,

재고 관리 시스템에서 다음과 같은 발주 요청이 발생했습니다.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
발주 정보
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
품목명: {item_name}
현재 재고: {current_stock}
최소 재고 기준: {min_stock}
발주 수량: {order_quantity}
발주 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

위 품목의 발주를 요청드립니다.
빠른 시일 내에 배송 부탁드립니다.

감사합니다.
        """
        
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # 이메일 발송
        if EMAIL_PASSWORD:
            server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.send_message(msg)
            server.quit()
            return True
        else:
            print(f"[이메일 미발송] 비밀번호 미설정: {item_name} -> {supplier_email}")
            return False
    except Exception as e:
        print(f"이메일 발송 오류: {e}")
        return False

@app.route('/send_notification_email/<int:item_id>', methods=['POST'])
@login_required
def send_notification_email(item_id):
    """담당자 알림 이메일 발송"""
    data = request.get_json()
    supplier_email = data.get('supplier_email', '')
    notification_message = data.get('notification_message', '')
    
    if not supplier_email:
        return jsonify({'success': False, 'message': '거래처이메일이 설정되지 않았습니다.'}), 400
    
    conn = get_db_connection()
    item = conn.execute('SELECT * FROM inventory WHERE id = ?', (item_id,)).fetchone()
    conn.close()
    
    if not item:
        return jsonify({'success': False, 'message': '항목을 찾을 수 없습니다.'}), 404
    
    # 담당자알림메시지가 없으면 기본 메시지 사용
    if not notification_message or notification_message.strip() == '':
        try:
            shortage_qty = item['shortage_quantity'] if item['shortage_quantity'] is not None else (item['min_stock'] - item['current_stock'] if item['current_stock'] < item['min_stock'] else 0)
        except (KeyError, IndexError):
            shortage_qty = item['min_stock'] - item['current_stock'] if item['current_stock'] < item['min_stock'] else 0
        try:
            reorder_qty = item['reorder_quantity'] if item['reorder_quantity'] is not None else 0
        except (KeyError, IndexError):
            reorder_qty = 0
        try:
            item_status = item['status'] if item['status'] else '정상'
        except (KeyError, IndexError):
            item_status = '정상'
        notification_message = f"""
재고 관리 시스템 알림

품목명: {item['item_name']}
현재 재고: {item['current_stock']}
안전재고: {item['min_stock']}
부족수량: {shortage_qty}
발주권장수량: {reorder_qty}
상태: {item_status}

알림 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
    
    try:
        # 이메일 내용 생성
        msg = MIMEMultipart()
        msg['From'] = EMAIL_SENDER
        msg['To'] = supplier_email
        msg['Subject'] = f'[재고관리알림] {item["item_name"]}'
        
        msg.attach(MIMEText(notification_message.strip(), 'plain', 'utf-8'))
        
        # 이메일 발송
        if EMAIL_PASSWORD:
            server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.send_message(msg)
            server.quit()
            
            # 발주 내역 저장
            conn = get_db_connection()
            conn.execute('''
                INSERT INTO orders (item_id, item_name, order_quantity, email_sent, email_sent_at)
                VALUES (?, ?, ?, 1, ?)
            ''', (item_id, item['item_name'], 0, datetime.now()))
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True,
                'message': f'{item["item_name"]}의 담당자알림메시지가 {supplier_email}로 발송되었습니다.'
            })
        else:
            return jsonify({
                'success': False,
                'message': '이메일 비밀번호가 설정되지 않았습니다.'
            }), 400
    except Exception as e:
        error_msg = str(e)
        print(f"이메일 발송 오류: {error_msg}")
        return jsonify({
            'success': False,
            'message': f'이메일 발송 중 오류가 발생했습니다: {error_msg}'
        }), 500

@app.route('/delete/<int:item_id>', methods=['POST'])
@login_required
def delete_item(item_id):
    """재고 항목 삭제"""
    conn = get_db_connection()
    conn.execute('DELETE FROM inventory WHERE id = ?', (item_id,))
    conn.commit()
    conn.close()
    flash('항목이 삭제되었습니다.', 'success')
    return redirect(url_for('index'))

@app.route('/orders')
@login_required
def orders():
    """발주 내역 조회"""
    conn = get_db_connection()
    orders_list = conn.execute('''
        SELECT * FROM orders ORDER BY order_date DESC
    ''').fetchall()
    conn.close()
    return render_template('orders.html', orders=orders_list)

@app.route('/upload', methods=['GET', 'POST'])
@login_required
def upload_excel():
    """엑셀 파일 업로드"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('파일이 선택되지 않았습니다.', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('파일이 선택되지 않았습니다.', 'error')
            return redirect(request.url)
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                df = pd.read_excel(file)
                
                # 엑셀 컬럼 매핑 (컬럼명에 따라 조정 필요)
                required_columns = ['품목명', '현재재고', '최소재고']
                if not all(col in df.columns for col in required_columns):
                    flash('엑셀 파일 형식이 올바르지 않습니다. (필수 컬럼: 품목명, 현재재고, 최소재고)', 'error')
                    return redirect(request.url)
                
                conn = get_db_connection()
                count = 0
                for _, row in df.iterrows():
                    item_name = str(row['품목명'])
                    current_stock = int(row['현재재고'])
                    min_stock = int(row['최소재고'])
                    unit = str(row.get('단위', '개'))
                    supplier_email = str(row.get('공급업체이메일', ''))
                    supplier_name = str(row.get('공급업체명', ''))
                    
                    # 기존 항목 확인
                    existing = conn.execute('SELECT id FROM inventory WHERE item_name = ?', (item_name,)).fetchone()
                    
                    if existing:
                        conn.execute('''
                            UPDATE inventory 
                            SET current_stock = ?, min_stock = ?, unit = ?, 
                                supplier_email = ?, supplier_name = ?, updated_at = ?
                            WHERE item_name = ?
                        ''', (current_stock, min_stock, unit, supplier_email, supplier_name, datetime.now(), item_name))
                    else:
                        conn.execute('''
                            INSERT INTO inventory (item_name, current_stock, min_stock, unit, supplier_email, supplier_name)
                            VALUES (?, ?, ?, ?, ?, ?)
                        ''', (item_name, current_stock, min_stock, unit, supplier_email, supplier_name))
                    
                    # 재고 부족 확인
                    if current_stock <= min_stock and supplier_email:
                        send_order_email(item_name, current_stock, min_stock, supplier_email, supplier_name)
                        conn.execute('''
                            INSERT INTO orders (item_id, item_name, order_quantity, email_sent, email_sent_at)
                            VALUES ((SELECT id FROM inventory WHERE item_name = ?), ?, ?, 1, ?)
                        ''', (item_name, item_name, min_stock * 3 - current_stock, datetime.now()))
                    
                    count += 1
                
                conn.commit()
                conn.close()
                flash(f'{count}개의 항목이 업로드되었습니다.', 'success')
            except Exception as e:
                flash(f'엑셀 파일 처리 중 오류가 발생했습니다: {str(e)}', 'error')
        
        return redirect(url_for('index'))
    
    return render_template('upload.html')

@app.route('/download')
@login_required
def download_excel():
    """재고 현황 엑셀 다운로드"""
    conn = get_db_connection()
    items = conn.execute('SELECT * FROM inventory').fetchall()
    conn.close()
    
    # 데이터프레임 생성
    data = []
    for item in items:
        data.append({
            '품목명': item['item_name'],
            '현재재고': item['current_stock'],
            '최소재고': item['min_stock'],
            '단위': item['unit'],
            '공급업체명': item['supplier_name'],
            '공급업체이메일': item['supplier_email'],
            '상태': '부족' if item['current_stock'] <= item['min_stock'] else '정상'
        })
    
    df = pd.DataFrame(data)
    
    # 엑셀 파일로 저장
    filename = f'inventory_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join('uploads', filename)
    os.makedirs('uploads', exist_ok=True)
    df.to_excel(filepath, index=False, engine='openpyxl')
    
    return send_file(filepath, as_attachment=True, download_name=filename)

# Vercel 배포를 위한 초기화는 api/index.py에서 처리
# 여기서는 로컬 실행 시에만 초기화

if __name__ == '__main__':
    init_db()
    load_excel_data()  # 앱 시작 시 엑셀 파일에서 기본 데이터 자동 로드
    os.makedirs('uploads', exist_ok=True)
    app.run(debug=True, host='0.0.0.0', port=5000)
